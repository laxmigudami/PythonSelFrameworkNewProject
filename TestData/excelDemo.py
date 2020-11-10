import openpyxl
# to write the data into file we have to create workbook object
# if we want to read the data from existing file then we have to use load_workbook() object
# to get active sheet wb.active
# to get value .value we have to use
#
from openpyxl import load_workbook
from openpyxl import workbook

book = openpyxl.load_workbook("C:\\datadriven\\pythonDemo1.xlsx")
sheet = book.active
dict = {}
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
# sheet.cell(row=2, column=2).value = "laxmi"
# print(sheet.cell(row=2, column=2).value)
#
# print(sheet.max_column)
# print(sheet.max_row)
# print(sheet["A5"].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column + 1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)


def generic_read_data(filename, sheetname, row, col):
    l_wb = load_workbook(filename)
    sheet = l_wb[sheetname]
    # l_wb.get_sheet_by_name
    res = sheet.cell(row=row, col=col).value

    return res


def max_rowCount(filename, sheetname):
    l_wb = load_workbook(filename)
    sheet = l_wb[sheetname]
    return sheet.max_row


def max_colCount(filename, sheetname):
    l_wb = load_workbook(filename)
    sheet = l_wb[sheetname]
    return sheet.max_column


# in any other files
# import exceldemo
# max_rows = exceldemo.max_rowCount("example.xlsx", "Sheet")
# max_col = exceldemo.max_colCount("example.xlsx", "Sheet")
# for i in range(1, max_rows+1):
#     username= exceldemo.generic_read_data("example.xlsx", "sheet", i,1)
#     password = exceldemo.generic_read_data("example.xlsx", "sheet", i, 2)
#     driver.find_element(xpath).sen_keys(username)
#     driver.find_element(xpath).send_keys(password)



