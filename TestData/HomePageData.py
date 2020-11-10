import openpyxl


class HomePageData:
    test_homepage_data = [
        {"firstname": "Laxmi", "Email": "gudamilaxmi444@gmail.com", "password": "laxmi@123", "gender": "Female"},
        {"firstname": "Preksha", "Email": "preksha@gmail.com", "password": "preksha@123", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):  # self parameter is required only when the method is declared as non static

        book = openpyxl.load_workbook("C:\\datadriven\\pythonDemo1.xlsx")
        sheet = book.active
        dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(dict)
        return [dict]
